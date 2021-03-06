from django.shortcuts import render,redirect,HttpResponse
from basicapp.models import profile,friends,Posts,Comments
from basicapp.forms import user_form,newpost,comment,profile_update
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.core.mail import send_mail
from openpyxl import Workbook

# Create your views here.
def registration(request):
    if request.method == 'POST':
        form_obj = user_form(request.POST)
        print('check1')
        if form_obj.is_valid():
            print('check2')
            username = form_obj.cleaned_data['username']
            firstname = form_obj.cleaned_data['first_name']
            lastname = form_obj.cleaned_data['last_name']
            pro_site = form_obj.cleaned_data['site']
            #pro_pic = form_obj.cleaned_data['picture']
           # birthdate = form_obj.cleaned_data['birthdate']
            password = form_obj.cleaned_data['password']
            user = User(username=username,first_name=firstname,last_name=lastname)#,birthdate=birthdate)
            user.set_password(password)
            user.save()
            profile_obj = profile(user=user,portfolio_site=pro_site)#,profile_pic=pro_pic)
            #user.save()
            print(user.username)
            profile_obj.save()
            friends.objects.create(tag=user)
            friends.objects.get(tag=user).save()
            return redirect('/login/')
            #return HttpResponse('done')
    else:
        form_obj= user_form()
        return render(request,'registration.html',{'form_1':form_obj,})


def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request,username= username,password=password)

        if user:
            login(request,user)
            print(str((request.user).first_name))
            email = request.user.email
            send_mail('login','Somebody Just Logged In!','jayant1616@gmail.com',[email])
            return redirect('/newsfeed')
        else:
            return HttpResponse('invalid credentials')
    else:
        return render(request,'login.html',{})

def user_logout(request):
    logout(request)
    return redirect('/')

def home(request):
    return render(request,'home.html',{})

"""
def update_profile(request):
    user = request.user
    profile_updt = profile(user = user) #profile model object

    if request.method == 'POST':
        form = profile_update(request.POST)
        if form.is_valid():
            print('checking update')
            #user.email = form.cleaned_data['email']
            profile_updt.profile_pic =  request.FILE["img"]
            #form.save()
            print(profile.objects.get(user=User.objects.get(username='admin')).profile_pic)

            print(profile_updt.profile_pic)
            print(profile_updt.profile_pic.url)
            profile_updt.save()
            return redirect('/')

    else:
        form = profile_update()
        return render(request,'profile_UPDATE.html',{'form':form})"""
#the folowing is not functioning due to form failing validation, working on it

def newsfeed(request):
    print(request)
    print(request.user.first_name)
    main_user = request.user
    load_friends = friends.objects.get(tag=main_user).following.all()
    post_all=Posts.objects.none()#creating an empty queryset
    for friends_all in load_friends:# basicaaly assigns users to filter out friends

        post_all = post_all | Posts.objects.filter(for_user = friends_all)
    dict_post = {'post_all':post_all,}

    return render(request,'Posts.html',dict_post)

def add_post(request):
    post_model = Posts(for_user = request.user)
    if request.method=='POST':
        post_object=newpost(request.method)
       # for test in post_object.errors:
        #     print('t')
         #    print(post_object.errors[test])
        if  post_object.is_valid():

            print('check1')
            post_model.user_post=post_object.cleaned_data['post_1']
            post_model.save()
            return redirect('admin/')

    else:
        post_object = newpost()
        return render(request,'newpost.html',{'post':post_object,})

def comments(request,id):
    print(request)
    post_get = Posts.objects.get(id = id)
    comment_object = Comments(for_comment=post_get,user = request.user)
    if request.method == 'POST':
        form = comment(request.POST)
        if form.is_valid():
            print('checkcomment')
            comment_object.user_comment = form.cleaned_data['comment_add']
            comment_object.save()
            return redirect('all_comments',id = id)
    else:
        form = comment()
        return render(request,'comments.html',{'form':form,})

def all_comments(request,id):
    post = Posts.objects.get(id=id)
    all_comments = Posts.objects.get(id=id).comments.all()
    context = {'comments':all_comments}
    return render(request,'all_comments.html',context)

def user_profiles(request):
    print(request)
    all_users = User.objects.all()
    context = {'users': all_users}
    return render(request,'user_profiles.html',context)



def ind_profile(request,user_id):
    profile_user = User.objects.get(id = user_id) #user of profile
    follow_list = friends.objects.get(tag = request.user).following.all() # logged in user's folllowiing list
    #follower_list = user.followers.all()
    #users_qset = User.objects.none()
    try :
        prof_pic = profile.objects.get(user=profile_user).profile_pic.url

    except:
        prof_pic = '/media/profile_pics/default.png'


    """for all in follower_list:
        users_qset = users_qset|all.tag
        """
    if profile_user in follow_list: # check if looged in user is following profile
        status='Already following'
    else:
        status = 'Follow'
    values={'profile_pic':prof_pic,'user_id':user_id,'username':profile_user.username,'firstname1':profile_user.first_name,'lastname1':profile_user.last_name,'follow_list':follow_list,'status':status}
    return render(request,'show_user.html',values)

def add_following(request,id):
    user = request.user
    if User.objects.get(id=id) not in friends.objects.get(tag=user).following.all():
        friends.objects.get(tag= user).following.add(User.objects.get(id=id))
        friends.objects.get(tag= user).save()

        return redirect('ind_profile',user_id = id)
    else:
        return HttpResponse('already following')

def user_info(request):
    if not request.user.is_staff:
        return HttpResponse('You need to be staff member to see this information')
    else:
    #for file
        info = Workbook()
        user_sheet = info.active
        Users_list = User.objects.all()
        v=1
        for all in Users_list:
            user_sheet.cell(row=v,column=1).value = 'User'+str(v)
            user_sheet.cell(row=v,column=2).value = all.username
            user_sheet.cell(row=v,column=3).value = all.first_name
            user_sheet.cell(row=v,column=4).value = all.email
            #user_sheet.cell(row=v,column=5).value = profile.objects.get(user=all).profile_pic
            v=v+1
        info.save('user_report.xlxs')

        return HttpResponse('Check the report in Base Folder')
